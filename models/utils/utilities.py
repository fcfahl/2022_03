import os,shutil,tarfile,time,subprocess,requests
# import wrapt,xlrd,xlsxwriter
import csv,yaml
import pandas as pd
import netCDF4 as nc
from pandas_ods_reader import read_ods
from datetime import date, datetime, timedelta
from zipfile import ZipFile
from tqdm import tqdm
from osgeo import gdal, osr

# ==============================================================================
#                                       FUNCTIONS
# ==============================================================================

# BACKUP


def backup_scripts(path):
    """ Backup scrips at the first day initialization of the script """

    today = date.today().strftime("%Y-%m-%d")
    folder_name = path.tmp.joinpath('BK')
    file_name = path.tmp.joinpath(folder_name, f"BK_{today}.tar.gz")

    create_folder(folder_name)

    if not os.path.exists(file_name):
        create_tar(path.script, file_name)
        print(f"\n\t\t ... backup file \t -> \t{file_name}\n")

# FILES


def save_file(file_full_path, content):
    """ Save file and create the folders, if needed  """

    os.makedirs(os.path.dirname(file_full_path), exist_ok=True)
    with open(file_full_path, "w", newline='\r\n') as f:
        f.write(content)

def save_list_to_file(filename, list):

    with open(filename, "w", newline='\r\n') as f:
        f.write('\n'.join(str(line) for line in list))


def delete_file(path):
    """ Delete file """

    if os.path.exists(path):
        os.remove(path)

def delete_files_by_extension(path, ext):
    """ Delete files by extension """
    for filename in os.listdir(path):
        if filename.endswith(ext):
            delete_file(os.path.join(path, filename))

def get_filename_from_path(path):
    """ extract the file name from a given path without extension"""

    return(os.path.splitext(os.path.basename(path))[0])


def copy_file(src, dest):
    """ copy files from src to dest """

    os.makedirs(os.path.dirname(dest), exist_ok=True)
    shutil.copyfile(src, dest)



# FOLDERS


def create_folder(path):
    """ Create new folder if it does not exist """

    if not os.path.exists(path):
        os.makedirs(path)


def delete_folder(path):
    """ Delete folder """

    try:
        shutil.rmtree(path)
    except OSError as e:
        print(f"Error: {e.filename} - {e.strerror}")

# COMPRESSION


def create_tar(path, file_name):
    """ Archive files as *.tar """

    with tarfile.open(file_name, mode='w:gz') as archive:
        archive.add(path, recursive=True)


def extract_tar_all(tar_file, path):
    """ Extract all files in a *.tar file """

    create_folder(path)

    tar = tarfile.open(tar_file)
    tar.extractall(path)
    tar.close()


def extract_tar_file(tar_file, file_name, path):
    """ Extract specific file in a *.tar file """

    create_folder(path)

    tar = tarfile.open(tar_file, 'r')
    for member in tar.getmembers():
        if file_name in member.name:
            # print ("\n_____________", file_name, path)
            tar.extract(member, path)


def extract_zip(zip_full_path, dst_path, overwrite=True):
    """ Decompress *.zip file """

    if not os.path.exists(dst_path) or overwrite == True:

        create_folder(dst_path)

        with ZipFile(zip_full_path, 'r') as zipObj:
            zipObj.extractall(dst_path)


def extract_zip_specific_file(zip_file, in_file, out_file):
    """ Decompress specific zip file """

    with ZipFile(zip_file, 'r') as zipObj:
        # print (zipObj.namelist()) # print list of files
        with open(out_file, 'wb') as f:
            f.write(zipObj.read(in_file))

################ CSV, ODS, TXT

def loat_txt_file(filename):
    
    with open(filename) as f:
        lines = f.readlines()

    return lines

def save_to_txt_file(file_full_path, content):
    """ Save txt into txt file  """

    file = open(file_full_path, 'w')
    file.write(content)
    file.close()


def load_ods(ods_file, sheet_name):
    """ Load a ods sheet based on its name """

    return (read_ods(ods_file, sheet_name))


def list_to_csv(list, csv_file, quote=csv.QUOTE_NONE):
    """ Export list as csv"""

    with open(csv_file, 'w', newline='') as f:
        wr = csv.writer(f, quoting=quote)
        wr.writerow(list)


def xlsx_to_csv(path, xlsx, sheetname=None):
    """ Convert xlsx file to csv """

    excel_file = os.path.join(path, xlsx)
    file_name = os.path.splitext(excel_file)[0]
    if sheetname:
        csv_file = os.path.join(path, f"{file_name}_{sheetname}.csv")
    else:
        csv_file = os.path.join(path, f"{file_name}.csv")

    print(csv_file)


def prepend_multiple_lines(file_name, list):
    """Insert given list of strings as a new lines at the beginning of a file"""
    # define name of temporary dummy file
    dummy_file = file_name + '.bak'
    # open given original file in read mode and dummy file in write mode
    with open(file_name, 'r') as read_obj, open(dummy_file, 'w') as write_obj:
        # Iterate over the given list of strings and write them to dummy file as lines
        for line in list:
            write_obj.write(line)
        # Read lines from original file one by one and append them to the dummy file
        for line in read_obj:
            write_obj.write(line)
    # remove original file
    os.remove(file_name)
    # Rename dummy file as the original file
    os.rename(dummy_file, file_name)

# YAML

def load_yaml(file_name):

    with open(file_name, 'r') as stream:
        try:
            yaml_content = yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(f'ERROR loading yaml file {file_name}: {exc}')

    return yaml_content


def save_yaml(file_name, dictionay):

    try:
        with open(file_name, 'w') as yamlfile:
            yaml.dump(dictionay, yamlfile, default_flow_style=False)
    except:
        print(f'NOT POSSIBLE TO EXPORT FILE -> {file_name}')


def append_yaml(file_name, dictionay):

    with open(file_name, 'r') as yamlfile:
        current_yaml = yaml.safe_load(yamlfile)  # Note the safe_load
        current_yaml.update(dictionay)

    if current_yaml:
        with open(file_name, 'w') as yamlfile:
            yaml.safe_dump(current_yaml, yamlfile)  # Also note the safe_dump


# XLS
# xlsxwriter does not allow editing xls files aftewards - not useful for multi editing

# def create_xls(path=None, file_name=None, sheet_name='sheet1', row=None, col=None, text=None, close=False):

#     if path:
#         # file = f'{path}/{file_name}'
#         file_ = path.joinpath(file_name)
#     else:
#         file_ = file_name

#     workbook = xlsxwriter.Workbook(file_)
#     worksheet = workbook.add_worksheet(sheet_name)

#     if text:
#         worksheet.write(row, col, text)


#     if close:   
#         workbook.close()


def create_xls(path=None, file_name=None, sheet_name='sheet1'):
    """ openpyxl allows later edition - xlsxwriter does not allow it """
    from openpyxl import Workbook

    if path:
        # file = f'{path}/{file_name}'
        file_ = path.joinpath(file_name)
    else:
        file_ = file_name

    wb = Workbook()       
    ws = wb.active
    ws.title = sheet_name  
    wb.save(file_)


def add_content_to_xls(path=None, file_name=None, sheet_name='sheet1', row=None, col=None, value=None, fgcolor=None):
    """ openpyxl allows later edition - xlsxwriter does not allow it """
    from openpyxl import load_workbook

    if path:
        # file = f'{path}/{file_name}'
        file_ = path.joinpath(file_name)
    else:
        file_ = file_name

    wb = load_workbook(file_)    
    ws = wb.get_sheet_by_name(sheet_name)
    cell = ws.cell(row,col)
    cell.value = value

    # if fgcolor:
    #     cell.fill = fgcolor

    wb.save(file_)

def set_xls_styles(path=None, file_name=None, sheet_name='sheet1', font=None, fill=None, alignment=None, row=None, col=None, gridlines=False, zoomscale = 70, fgcolor=None):


    def apply_styles(cell):

        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
            

    from openpyxl import load_workbook

    if path:
        # file = f'{path}/{file_name}'
        file_ = path.joinpath(file_name)
    else:
        file_ = file_name

    wb = load_workbook(file_)    
    ws = wb.get_sheet_by_name(sheet_name)

    ws.sheet_view.showGridLines = gridlines
    ws.sheet_view.zoomScale = zoomscale

    for row_xls in ws.iter_rows():
        for cell_xls in row_xls:

            if not row and not col: # apply to all cells
                cell = cell_xls
                
            else: # apply to specific cells
                cell = ws.cell(row,col)    

            apply_styles(cell)

    wb.save(file_)
        

# PANDAS


def csv_to_pd(csv, path, encoding="utf-8"):  # utf-8
    """ Load csv into pandas dataframe """

    return (pd.read_csv(f"{path}/{csv}", encoding=encoding))

# DOWNLOAD


def download_file(url, file_name, overwrite=True):
    """ Download files from web pages """

    if not os.path.exists(file_name) or overwrite == True:
        web_file = requests.get(url)
        open(file_name, 'wb').write(web_file.content)


def download_file_progress_bar(url, file_name, overwrite=True):
    """ 
        Download files from web pages with progress bar
        https://newbedev.com/progress-bar-while-download-file-over-http-with-requests
    """

    if not os.path.exists(file_name) or overwrite == True:

        response = requests.get(url, stream=True)
        total_size_in_bytes = int(response.headers.get('content-length', 0))
        block_size = 1024  # 1 Kibibyte
        progress_bar = tqdm(total=total_size_in_bytes,
                            unit='iB', unit_scale=True)
        with open(file_name, 'wb') as file:
            for data in response.iter_content(block_size):
                progress_bar.update(len(data))
                file.write(data)
        progress_bar.close()
        if total_size_in_bytes != 0 and progress_bar.n != total_size_in_bytes:
            print("ERROR on DOWLOADING, something went wrong")

# BASH


def run_bash_script(path, script):
    """ Run bash script by calling os.subprocess """

    os.chdir(path)  # change path to the source data - otherwise gets error
    subprocess.call(f'./{script}', shell=True)


def run_bash_reproject(path, script, in_file, out_file, in_proj, out_proj):
    """ Run bash script to reproject raster lauers by calling os.subprocess """

    process = f"sh {path}{script} {in_file} {out_file} {in_proj} '{out_proj}'"

    os.chdir(path)  # change path to the source data - otherwise gets error
    # subprocess.call([process], shell=True)

    subprocess.Popen([process], shell=True)


def run_bash_reproject_bbox(path, script, in_file, out_file, in_proj, out_proj, xmin=None, ymin=None, xmax=None, ymax=None, bbox=None):
    """ Run bash script to reproject raster lauers by calling os.subprocess """

    if bbox:
        process = f"sh {script} {in_file} {out_file} {in_proj} '{out_proj}' {bbox} "
    else:
        process = f"sh {script} {in_file} {out_file} {in_proj} '{out_proj}' {xmin} {ymin} {xmax} {ymax} "

    print (process)


    os.chdir(path)  # change path to the source data - otherwise gets error
    # subprocess.call([process], shell=True)
    subprocess.Popen([process], shell=True)

# TIME


def get_time():
    """ Get the current time """
    return time.time()


def get_total_time(start, end):
    """ Calculate the total time between two dates """
    return round((end - start) / 60, 2)


def get_hours_in_year(year):

    date_format = "%Y-%m-%d %H:%M:%S"
    start_date_time = datetime.strptime(f"{year}-01-01 00:00:00", date_format)
    end_date_time = datetime.strptime(f"{year+1}-01-01 00:00:00", date_format)

    interval = end_date_time - start_date_time

    return divmod(interval.total_seconds(), 3600)[0]

# GDAL


def reproject_raster(in_file, out_file, in_proj, out_proj):
    """ Reproject files using GDAL - provide full path for input and output files  
    https://gis.stackexchange.com/questions/233589/re-project-raster-in-python-using-gdal """

    input_raster = gdal.Open(in_file)
    warp = gdal.Warp(out_file, input_raster, srcSRS=in_proj, dstSRS=out_proj)
    warp = None  # Closes the files


def reproject_netcdf(in_file, band, out_file, in_proj=4326, out_proj=3035, bbox=None, res=None ):
    """ Reproject files using GDAL - provide full path for input and output files  
    https://gis.stackexchange.com/questions/379114/using-gdal-warp-on-netcdf-file-simple-case-fails """

    srs = osr.SpatialReference()
    dst = osr.SpatialReference()    

    srs.ImportFromEPSG(in_proj)
    dst.ImportFromEPSG(out_proj)

    kwargs = dict(
        format       = 'GTiff',
        copyMetadata = False,
        outputBounds = bbox, # BBOX in the target SRS
        srcSRS       = srs,
        dstSRS       = dst,
        xRes         = res, # output resolution in target SRS
        yRes         = res, # output resolution in target SRS

    )

    input_raster = gdal.Open(f"NETCDF:{in_file}:{band}")
    warp = gdal.Warp(out_file, input_raster, **kwargs)
    warp = None # Closes the files

# NETCDF

def get_netcdf_metadata(netcdf_file, variable):
    """ get metadata from netcdf file """

    ds = nc.Dataset(netcdf_file)

    scale = ds.scale_factor
    offset = ds.add_offset

    return ds.variables[variable]


# GENERIC

def get_next_letter(letter):

    if letter == "Z": # If Z encountered change to A
        next_letter = chr(ord(letter)-25)
    else:
        next_letter = chr(ord(letter) + 1)

    return next_letter


################### Script control

def pause():
    input("Press any key to continue...")
